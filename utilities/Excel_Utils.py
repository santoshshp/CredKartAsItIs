import openpyxl


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def read_data(file, sheet_name, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=column_num).value

def write_data(file, sheet_name, row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=column_num).value = data
    workbook.save(file)